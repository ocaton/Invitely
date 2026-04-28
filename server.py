from __future__ import annotations

import json
import os
import re
import secrets
import threading
import time
import uuid
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("INVITELY_DATA_DIR", str(BASE_DIR / "data")))
WORKBOOK_PATH = DATA_DIR / "gatherly-data.xlsx"
GOOGLE_CONNECTION_PATH = DATA_DIR / "google_sheets_connection.json"
ADMIN_PASSWORD = os.getenv("INVITELY_ADMIN_PASSWORD", "m3-host-2026")
SERVER_HOST = os.getenv("HOST", "127.0.0.1")
SERVER_PORT = int(os.getenv("PORT", "8000"))
SESSION_COOKIE_NAME = "invitely_admin_session"
SESSION_TTL_SECONDS = 60 * 60 * 24 * 7
GOOGLE_SCOPES = [
    "openid",
    "email",
    "profile",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

EVENT_HEADERS = [
    "id",
    "title",
    "subtitle",
    "startDate",
    "startTime",
    "endTime",
    "venue",
    "town",
    "address",
    "category",
    "host",
    "hostImageUrl",
    "description",
    "featured",
    "emoji",
    "capacity",
    "timezones",
    "eventImageUrl",
    "sourceUrl",
]

RSVP_HEADERS = [
    "createdAt",
    "eventId",
    "eventTitle",
    "name",
    "email",
    "phone",
    "guestCount",
    "notes",
    "response",
]

SETTINGS_HEADERS = ["key", "value"]
DEFAULT_SETTINGS = {
    "appName": "Invitely",
    "logoMark": "I",
}

SEEDED_EVENT_ID = "mothers-day-2026"
SEEDED_EVENT = {
    "id": SEEDED_EVENT_ID,
    "title": "Mother's Day Lunch",
    "subtitle": "A special lunch to honor mothers, grandmothers, and aunts.",
    "startDate": "2026-05-10",
    "startTime": "12:00",
    "endTime": "",
    "venue": "RSVP to see location",
    "town": "",
    "address": "",
    "category": "Lunch",
    "host": "Milt Lauenstein",
    "hostImageUrl": "",
    "description": (
        "Join us for a special Mother's Day Lunch as we celebrate and honor all the "
        "wonderful mothers, grandmothers, and aunts who have made our lives meaningful. "
        "Enjoy some tasty morsels, warm company, and a relaxing afternoon together."
    ),
    "featured": True,
    "emoji": "MD",
    "capacity": "",
    "timezones": "EST,IST",
    "eventImageUrl": "",
    "sourceUrl": "https://partiful.com/e/5Y7B9wquDfUuVZIMevLu?",
}

LEGACY_DEFAULT_TITLE = "Mother's Day Gathering"
STATIC_ROUTES = {
    "/": "index.html",
    "/index.html": "index.html",
    "/styles.css": "styles.css",
    "/app.js": "app.js",
}

WORKBOOK_LOCK = threading.Lock()
ADMIN_SESSIONS: dict[str, dict[str, object]] = {}


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or uuid.uuid4().hex[:8]


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", start_color="6750A4", end_color="6750A4")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def ensure_headers(sheet, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = header
        header_style(cell)


def event_to_row(event: dict[str, object]) -> list[object]:
    return [event.get(header, "") for header in EVENT_HEADERS]


def sheet_rows_from_actual_headers(sheet, target_headers: list[str]) -> list[dict[str, object]]:
    actual_headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        actual_record = {}
        for index, header in enumerate(actual_headers):
            if not header:
                continue
            actual_record[header] = row[index] if index < len(row) and row[index] is not None else ""
        rows.append({header: actual_record.get(header, "") for header in target_headers})
    return rows


def rewrite_sheet_rows(sheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    max_rows = sheet.max_row
    if max_rows > 1:
        sheet.delete_rows(2, max_rows - 1)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def upsert_seed_event(sheet) -> None:
    row_index = None
    legacy_row_index = None

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        row_id = str(row[0].value or "").strip()
        row_title = str(row[1].value or "").strip()
        if row_id == SEEDED_EVENT_ID:
            row_index = index
            break
        if row_title == LEGACY_DEFAULT_TITLE:
            legacy_row_index = index

    target_index = row_index or legacy_row_index

    if target_index is None:
        if sheet.max_row == 1:
            sheet.append(event_to_row(SEEDED_EVENT))
        return

    if sheet.max_row != 2:
        return

    for row in sheet.iter_rows(min_row=2, values_only=False):
        row[11].value = False

    for column_index, value in enumerate(event_to_row(SEEDED_EVENT), start=1):
        sheet.cell(row=target_index, column=column_index, value=value)


def ensure_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if not WORKBOOK_PATH.exists():
        workbook = Workbook()
        events_sheet = workbook.active
        events_sheet.title = "Events"
        events_sheet.append(EVENT_HEADERS)
        ensure_headers(events_sheet, EVENT_HEADERS)
        events_sheet.append(event_to_row(SEEDED_EVENT))
        events_sheet.freeze_panes = "A2"

        rsvp_sheet = workbook.create_sheet("RSVPs")
        rsvp_sheet.append(RSVP_HEADERS)
        ensure_headers(rsvp_sheet, RSVP_HEADERS)
        rsvp_sheet.freeze_panes = "A2"

        settings_sheet = workbook.create_sheet("Settings")
        settings_sheet.append(SETTINGS_HEADERS)
        ensure_headers(settings_sheet, SETTINGS_HEADERS)
        for key, value in DEFAULT_SETTINGS.items():
            settings_sheet.append([key, value])
        settings_sheet.freeze_panes = "A2"

        workbook.save(WORKBOOK_PATH)
        workbook.close()
        return

    workbook = load_workbook(WORKBOOK_PATH)

    events_sheet = workbook["Events"] if "Events" in workbook.sheetnames else workbook.create_sheet("Events", 0)
    event_rows = sheet_rows_from_actual_headers(events_sheet, EVENT_HEADERS) if events_sheet.max_row >= 1 else []
    if events_sheet.max_row == 0:
        events_sheet.append(EVENT_HEADERS)
    ensure_headers(events_sheet, EVENT_HEADERS)
    rewrite_sheet_rows(events_sheet, EVENT_HEADERS, event_rows)
    upsert_seed_event(events_sheet)
    events_sheet.freeze_panes = "A2"

    rsvp_sheet = workbook["RSVPs"] if "RSVPs" in workbook.sheetnames else workbook.create_sheet("RSVPs")
    rsvp_rows = sheet_rows_from_actual_headers(rsvp_sheet, RSVP_HEADERS) if rsvp_sheet.max_row >= 1 else []
    if rsvp_sheet.max_row == 0:
        rsvp_sheet.append(RSVP_HEADERS)
    ensure_headers(rsvp_sheet, RSVP_HEADERS)
    rewrite_sheet_rows(rsvp_sheet, RSVP_HEADERS, rsvp_rows)
    rsvp_sheet.freeze_panes = "A2"

    settings_sheet = workbook["Settings"] if "Settings" in workbook.sheetnames else workbook.create_sheet("Settings")
    settings_rows = (
        sheet_rows_from_actual_headers(settings_sheet, SETTINGS_HEADERS) if settings_sheet.max_row >= 1 else []
    )
    if settings_sheet.max_row == 0:
        settings_sheet.append(SETTINGS_HEADERS)
    ensure_headers(settings_sheet, SETTINGS_HEADERS)
    rewrite_sheet_rows(settings_sheet, SETTINGS_HEADERS, settings_rows)
    existing_keys = {
        str(settings_sheet.cell(row=index, column=1).value or "").strip(): index
        for index in range(2, settings_sheet.max_row + 1)
    }
    for key, value in DEFAULT_SETTINGS.items():
        if key not in existing_keys:
            settings_sheet.append([key, value])
    settings_sheet.freeze_panes = "A2"

    workbook.save(WORKBOOK_PATH)
    workbook.close()


def read_sheet_rows(sheet_name: str, headers: list[str]) -> list[dict[str, object]]:
    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook[sheet_name]
        rows = sheet_rows_from_actual_headers(sheet, headers)
        workbook.close()
    return rows


def workbook_get_events() -> list[dict[str, object]]:
    rows = read_sheet_rows("Events", EVENT_HEADERS)
    events = []
    for row in rows:
        events.append(
            {
                "id": str(row["id"]),
                "title": str(row["title"]),
                "subtitle": str(row["subtitle"]),
                "startDate": str(row["startDate"]),
                "startTime": str(row["startTime"]),
                "endTime": str(row["endTime"]),
                "venue": str(row["venue"]),
                "town": str(row["town"]),
                "address": str(row["address"]),
                "category": str(row["category"]),
                "host": str(row["host"]),
                "hostImageUrl": str(row["hostImageUrl"]),
                "description": str(row["description"]),
                "featured": parse_bool(row["featured"]),
                "emoji": str(row["emoji"]),
                "capacity": str(row["capacity"]),
                "timezones": str(row["timezones"]),
                "eventImageUrl": str(row["eventImageUrl"]),
                "sourceUrl": str(row["sourceUrl"]),
            }
        )
    return sorted(events, key=lambda item: (not item["featured"], item["startDate"], item["startTime"]))


def workbook_get_rsvps() -> list[dict[str, object]]:
    rows = read_sheet_rows("RSVPs", RSVP_HEADERS)
    rsvps = [
        {
            "createdAt": str(row["createdAt"]),
            "eventId": str(row["eventId"]),
            "eventTitle": str(row["eventTitle"]),
            "name": str(row["name"]),
            "email": str(row["email"]),
            "phone": str(row["phone"]),
            "guestCount": str(row["guestCount"]),
            "notes": str(row["notes"]),
            "response": str(row["response"] or "Declined"),
        }
        for row in rows
    ]
    rsvps.reverse()
    return rsvps


def workbook_get_settings() -> dict[str, str]:
    rows = read_sheet_rows("Settings", SETTINGS_HEADERS)
    settings = DEFAULT_SETTINGS.copy()
    for row in rows:
        key = str(row["key"]).strip()
        value = str(row["value"]).strip()
        if key:
            settings[key] = value
    return settings


def workbook_save_settings(payload: dict[str, object]) -> None:
    app_name = str(payload.get("appName", "")).strip()
    logo_mark = str(payload.get("logoMark", "")).strip()

    if not app_name:
        raise ValueError("App name is required.")
    if not logo_mark:
        raise ValueError("Logo mark is required.")

    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook["Settings"]
        current_rows = {
            str(sheet.cell(row=index, column=1).value or "").strip(): index
            for index in range(2, sheet.max_row + 1)
        }
        for key, value in {"appName": app_name, "logoMark": logo_mark}.items():
            row_index = current_rows.get(key)
            if row_index is None:
                sheet.append([key, value])
            else:
                sheet.cell(row=row_index, column=2, value=value)
        workbook.save(WORKBOOK_PATH)
        workbook.close()


def workbook_upsert_event(payload: dict[str, object]) -> None:
    title = str(payload.get("title", "")).strip()
    start_date = str(payload.get("startDate", "")).strip()
    event_id = str(payload.get("id", "")).strip() or slugify(f"{title}-{start_date}")

    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook["Events"]
        header_map = {
            str(sheet.cell(row=1, column=index).value or "").strip(): index
            for index in range(1, sheet.max_column + 1)
        }
        row_index = None

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if str(row[0].value or "").strip() == event_id:
                row_index = index
                break

        source_url = str(payload.get("sourceUrl", "")).strip()
        if row_index is not None and not source_url:
            source_column = header_map.get("sourceUrl")
            if source_column:
                source_url = str(sheet.cell(row=row_index, column=source_column).value or "").strip()

        row_values = [
            event_id,
            title,
            str(payload.get("subtitle", "")).strip(),
            start_date,
            str(payload.get("startTime", "")).strip(),
            str(payload.get("endTime", "")).strip(),
            str(payload.get("venue", "")).strip(),
            str(payload.get("town", "")).strip(),
            str(payload.get("address", "")).strip(),
            str(payload.get("category", "")).strip(),
            str(payload.get("host", "")).strip(),
            str(payload.get("hostImageUrl", "")).strip(),
            str(payload.get("description", "")).strip(),
            bool(payload.get("featured", False)),
            str(payload.get("emoji", "")).strip(),
            str(payload.get("capacity", "")).strip(),
            str(payload.get("timezones", "")).strip(),
            str(payload.get("eventImageUrl", "")).strip(),
            source_url,
        ]

        featured_column = header_map.get("featured", 14)
        if row_values[13]:
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row[featured_column - 1].value = False

        if row_index is None:
            sheet.append(row_values)
        else:
            for column_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(WORKBOOK_PATH)
        workbook.close()


def workbook_delete_event(event_id: str) -> None:
    if not event_id:
        raise ValueError("Event id is required.")

    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook["Events"]
        target_row = None

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if str(row[0].value or "").strip() == event_id:
                target_row = index
                break

        if target_row is None:
            workbook.close()
            raise ValueError("Event not found.")

        sheet.delete_rows(target_row, 1)
        workbook.save(WORKBOOK_PATH)
        workbook.close()


def workbook_add_rsvp(payload: dict[str, object]) -> None:
    event_id = str(payload.get("eventId", "")).strip()
    name = str(payload.get("name", "")).strip()
    email = str(payload.get("email", "")).strip()
    phone = str(payload.get("phone", "")).strip()
    guest_count = str(payload.get("guestCount", "")).strip()
    notes = str(payload.get("notes", "")).strip()
    response = str(payload.get("response", "")).strip()

    if not event_id:
        raise ValueError("Choose an event before submitting.")
    if not name:
        raise ValueError("Your full name is required.")
    if phone and not re.fullmatch(r"[+()\-\d\s]{7,20}", phone):
        raise ValueError("Please enter a valid phone number.")
    if not guest_count:
        raise ValueError("Please enter how many guests are coming.")
    if response not in {"Going", "Unsure", "Declined"}:
        raise ValueError("Pick Going, Unsure, or Decline before submitting.")

    event = next((item for item in workbook_get_events() if item["id"] == event_id), None)
    if event is None:
        raise ValueError("That event could not be found.")

    with WORKBOOK_LOCK:
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook["RSVPs"]
        sheet.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                event_id,
                event["title"],
                name,
                email,
                phone,
                guest_count,
                notes,
                response,
            ]
        )
        workbook.save(WORKBOOK_PATH)
        workbook.close()


def load_json_file(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text("utf-8"))
    except json.JSONDecodeError:
        return {}


def save_json_file(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), "utf-8")


def google_oauth_configured() -> bool:
    return bool(os.getenv("GOOGLE_CLIENT_ID") and os.getenv("GOOGLE_CLIENT_SECRET"))


def load_google_connection() -> dict[str, object]:
    return load_json_file(GOOGLE_CONNECTION_PATH)


def save_google_connection(payload: dict[str, object]) -> None:
    save_json_file(GOOGLE_CONNECTION_PATH, payload)


def google_connected() -> bool:
    payload = load_google_connection()
    return bool(payload.get("refresh_token") and payload.get("spreadsheet_id"))


def import_google_libraries():
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import Flow
    from googleapiclient.discovery import build

    return Request, Credentials, Flow, build


def build_external_url(handler: BaseHTTPRequestHandler, path: str) -> str:
    proto = handler.headers.get("X-Forwarded-Proto", "http")
    host = handler.headers.get("Host", f"{SERVER_HOST}:{SERVER_PORT}")
    return f"{proto}://{host}{path}"


def get_google_redirect_uri(handler: BaseHTTPRequestHandler) -> str:
    configured = os.getenv("GOOGLE_REDIRECT_URI", "").strip()
    if configured:
        return configured
    return build_external_url(handler, "/auth/google/callback")


def google_client_config(handler: BaseHTTPRequestHandler) -> dict[str, object]:
    return {
        "web": {
            "client_id": os.getenv("GOOGLE_CLIENT_ID", ""),
            "client_secret": os.getenv("GOOGLE_CLIENT_SECRET", ""),
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [get_google_redirect_uri(handler)],
        }
    }


def get_google_credentials() -> object:
    if not google_oauth_configured():
        raise RuntimeError("Google OAuth is not configured.")

    connection = load_google_connection()
    refresh_token = str(connection.get("refresh_token", "")).strip()
    if not refresh_token:
        raise RuntimeError("Google Sheets is not connected yet.")

    Request, Credentials, _, _ = import_google_libraries()
    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=os.getenv("GOOGLE_CLIENT_ID", ""),
        client_secret=os.getenv("GOOGLE_CLIENT_SECRET", ""),
        scopes=GOOGLE_SCOPES,
    )
    creds.refresh(Request())
    return creds


def google_service(name: str, version: str, credentials) -> object:
    _, _, _, build = import_google_libraries()
    return build(name, version, credentials=credentials, cache_discovery=False)


def google_sheet_titles(sheets_api, spreadsheet_id: str) -> list[str]:
    metadata = (
        sheets_api.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets.properties.title")
        .execute()
    )
    return [sheet["properties"]["title"] for sheet in metadata.get("sheets", [])]


def ensure_google_sheet_tabs(sheets_api, spreadsheet_id: str) -> None:
    wanted = ["Settings", "Events", "RSVPs"]
    existing = set(google_sheet_titles(sheets_api, spreadsheet_id))
    requests = []
    for title in wanted:
        if title not in existing:
            requests.append({"addSheet": {"properties": {"title": title}}})

    if requests:
        sheets_api.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body={"requests": requests}
        ).execute()


def google_write_table(spreadsheet_id: str, sheet_name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    credentials = get_google_credentials()
    sheets_api = google_service("sheets", "v4", credentials)
    ensure_google_sheet_tabs(sheets_api, spreadsheet_id)
    values = [headers]
    for row in rows:
        values.append([row.get(header, "") for header in headers])

    range_name = f"{sheet_name}!A1:Z"
    sheets_api.spreadsheets().values().clear(spreadsheetId=spreadsheet_id, range=range_name).execute()
    sheets_api.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!A1",
        valueInputOption="RAW",
        body={"values": values},
    ).execute()


def google_read_table(spreadsheet_id: str, sheet_name: str, headers: list[str]) -> list[dict[str, str]]:
    credentials = get_google_credentials()
    sheets_api = google_service("sheets", "v4", credentials)
    ensure_google_sheet_tabs(sheets_api, spreadsheet_id)
    response = (
        sheets_api.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}!A:Z")
        .execute()
    )
    values = response.get("values", [])
    if not values:
        return []

    actual_headers = [str(value).strip() for value in values[0]]
    rows = []
    for raw_row in values[1:]:
        if not any(cell not in ("", None) for cell in raw_row):
            continue
        actual_record = {}
        for index, header in enumerate(actual_headers):
            if not header:
                continue
            actual_record[header] = raw_row[index] if index < len(raw_row) else ""
        rows.append({header: actual_record.get(header, "") for header in headers})
    return rows


def google_create_spreadsheet(title: str, credentials) -> dict[str, str]:
    sheets_api = google_service("sheets", "v4", credentials)
    response = sheets_api.spreadsheets().create(
        body={
            "properties": {"title": title},
            "sheets": [{"properties": {"title": "Settings"}}, {"properties": {"title": "Events"}}, {"properties": {"title": "RSVPs"}}],
        }
    ).execute()
    return {
        "spreadsheet_id": response["spreadsheetId"],
        "spreadsheet_url": response["spreadsheetUrl"],
    }


def google_get_events() -> list[dict[str, object]]:
    connection = load_google_connection()
    rows = google_read_table(str(connection.get("spreadsheet_id", "")), "Events", EVENT_HEADERS)
    events = []
    for row in rows:
        events.append(
            {
                "id": str(row["id"]),
                "title": str(row["title"]),
                "subtitle": str(row["subtitle"]),
                "startDate": str(row["startDate"]),
                "startTime": str(row["startTime"]),
                "endTime": str(row["endTime"]),
                "venue": str(row["venue"]),
                "town": str(row["town"]),
                "address": str(row["address"]),
                "category": str(row["category"]),
                "host": str(row["host"]),
                "hostImageUrl": str(row["hostImageUrl"]),
                "description": str(row["description"]),
                "featured": parse_bool(row["featured"]),
                "emoji": str(row["emoji"]),
                "capacity": str(row["capacity"]),
                "timezones": str(row["timezones"]),
                "eventImageUrl": str(row["eventImageUrl"]),
                "sourceUrl": str(row["sourceUrl"]),
            }
        )
    return sorted(events, key=lambda item: (not item["featured"], item["startDate"], item["startTime"]))


def google_get_rsvps() -> list[dict[str, object]]:
    connection = load_google_connection()
    rows = google_read_table(str(connection.get("spreadsheet_id", "")), "RSVPs", RSVP_HEADERS)
    rsvps = [
        {
            "createdAt": str(row["createdAt"]),
            "eventId": str(row["eventId"]),
            "eventTitle": str(row["eventTitle"]),
            "name": str(row["name"]),
            "email": str(row["email"]),
            "phone": str(row["phone"]),
            "guestCount": str(row["guestCount"]),
            "notes": str(row["notes"]),
            "response": str(row["response"] or "Declined"),
        }
        for row in rows
    ]
    rsvps.reverse()
    return rsvps


def google_get_settings() -> dict[str, str]:
    connection = load_google_connection()
    rows = google_read_table(str(connection.get("spreadsheet_id", "")), "Settings", SETTINGS_HEADERS)
    settings = DEFAULT_SETTINGS.copy()
    for row in rows:
        key = str(row["key"]).strip()
        value = str(row["value"]).strip()
        if key:
            settings[key] = value
    return settings


def google_save_settings(payload: dict[str, object]) -> None:
    connection = load_google_connection()
    settings = google_get_settings()
    settings["appName"] = str(payload.get("appName", "")).strip()
    settings["logoMark"] = str(payload.get("logoMark", "")).strip()
    google_write_table(
        str(connection.get("spreadsheet_id", "")),
        "Settings",
        SETTINGS_HEADERS,
        [{"key": key, "value": value} for key, value in settings.items()],
    )


def google_upsert_event(payload: dict[str, object]) -> None:
    connection = load_google_connection()
    events = google_get_events()
    event_id = str(payload.get("id", "")).strip() or slugify(
        f"{str(payload.get('title', '')).strip()}-{str(payload.get('startDate', '')).strip()}"
    )

    updated_event = {
        "id": event_id,
        "title": str(payload.get("title", "")).strip(),
        "subtitle": str(payload.get("subtitle", "")).strip(),
        "startDate": str(payload.get("startDate", "")).strip(),
        "startTime": str(payload.get("startTime", "")).strip(),
        "endTime": str(payload.get("endTime", "")).strip(),
        "venue": str(payload.get("venue", "")).strip(),
        "town": str(payload.get("town", "")).strip(),
        "address": str(payload.get("address", "")).strip(),
        "category": str(payload.get("category", "")).strip(),
        "host": str(payload.get("host", "")).strip(),
        "hostImageUrl": str(payload.get("hostImageUrl", "")).strip(),
        "description": str(payload.get("description", "")).strip(),
        "featured": bool(payload.get("featured", False)),
        "emoji": str(payload.get("emoji", "")).strip(),
        "capacity": str(payload.get("capacity", "")).strip(),
        "timezones": str(payload.get("timezones", "")).strip(),
        "eventImageUrl": str(payload.get("eventImageUrl", "")).strip(),
        "sourceUrl": str(payload.get("sourceUrl", "")).strip(),
    }

    found = False
    for index, event in enumerate(events):
        if event["id"] == event_id:
            if not updated_event["sourceUrl"]:
                updated_event["sourceUrl"] = event.get("sourceUrl", "")
            events[index] = updated_event
            found = True
            break

    if not found:
        events.append(updated_event)

    if updated_event["featured"]:
        for event in events:
            event["featured"] = event["id"] == event_id

    google_write_table(str(connection.get("spreadsheet_id", "")), "Events", EVENT_HEADERS, events)


def google_delete_event(event_id: str) -> None:
    connection = load_google_connection()
    events = [event for event in google_get_events() if event["id"] != event_id]
    google_write_table(str(connection.get("spreadsheet_id", "")), "Events", EVENT_HEADERS, events)


def google_add_rsvp(payload: dict[str, object]) -> None:
    connection = load_google_connection()
    rsvps = google_get_rsvps()
    events = google_get_events()
    event = next((item for item in events if item["id"] == str(payload.get("eventId", "")).strip()), None)
    if event is None:
        raise ValueError("That event could not be found.")

    rsvps.insert(
        0,
        {
            "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "eventId": str(payload.get("eventId", "")).strip(),
            "eventTitle": event["title"],
            "name": str(payload.get("name", "")).strip(),
            "email": str(payload.get("email", "")).strip(),
            "phone": str(payload.get("phone", "")).strip(),
            "guestCount": str(payload.get("guestCount", "")).strip(),
            "notes": str(payload.get("notes", "")).strip(),
            "response": str(payload.get("response", "")).strip(),
        },
    )

    google_write_table(str(connection.get("spreadsheet_id", "")), "RSVPs", RSVP_HEADERS, list(reversed(rsvps)))


def migrate_workbook_to_google() -> None:
    connection = load_google_connection()
    spreadsheet_id = str(connection.get("spreadsheet_id", "")).strip()
    if not spreadsheet_id:
        raise RuntimeError("Google spreadsheet is not available for migration.")

    workbook_settings = workbook_get_settings()
    workbook_events = workbook_get_events()
    workbook_rsvps = list(reversed(workbook_get_rsvps()))

    google_write_table(
        spreadsheet_id,
        "Settings",
        SETTINGS_HEADERS,
        [{"key": key, "value": value} for key, value in workbook_settings.items()],
    )
    google_write_table(spreadsheet_id, "Events", EVENT_HEADERS, workbook_events)
    google_write_table(spreadsheet_id, "RSVPs", RSVP_HEADERS, workbook_rsvps)


def get_settings() -> dict[str, str]:
    if google_connected():
        try:
            return google_get_settings()
        except Exception:
            return workbook_get_settings()
    return workbook_get_settings()


def get_events() -> list[dict[str, object]]:
    if google_connected():
        try:
            return google_get_events()
        except Exception:
            return workbook_get_events()
    return workbook_get_events()


def get_rsvps() -> list[dict[str, object]]:
    if google_connected():
        try:
            return google_get_rsvps()
        except Exception:
            return workbook_get_rsvps()
    return workbook_get_rsvps()


def save_settings(payload: dict[str, object]) -> None:
    workbook_save_settings(payload)
    if google_connected():
        google_save_settings(payload)


def upsert_event(payload: dict[str, object]) -> None:
    workbook_upsert_event(payload)
    if google_connected():
        google_upsert_event(payload)


def delete_event(event_id: str) -> None:
    workbook_delete_event(event_id)
    if google_connected():
        google_delete_event(event_id)


def add_rsvp(payload: dict[str, object]) -> None:
    workbook_add_rsvp(payload)
    if google_connected():
        google_add_rsvp(payload)


def cleanup_expired_sessions() -> None:
    now = time.time()
    expired = [token for token, data in ADMIN_SESSIONS.items() if float(data.get("expires_at", 0)) <= now]
    for token in expired:
        ADMIN_SESSIONS.pop(token, None)


def create_admin_session() -> str:
    cleanup_expired_sessions()
    token = secrets.token_urlsafe(32)
    ADMIN_SESSIONS[token] = {"expires_at": time.time() + SESSION_TTL_SECONDS}
    return token


def get_session_token(handler: BaseHTTPRequestHandler) -> str:
    cookie_header = handler.headers.get("Cookie", "")
    if not cookie_header:
        return ""
    cookie = SimpleCookie()
    cookie.load(cookie_header)
    morsel = cookie.get(SESSION_COOKIE_NAME)
    return morsel.value if morsel else ""


def get_admin_session(handler: BaseHTTPRequestHandler) -> dict[str, object] | None:
    cleanup_expired_sessions()
    token = get_session_token(handler)
    if not token:
        return None
    session = ADMIN_SESSIONS.get(token)
    if not session:
        return None
    session["expires_at"] = time.time() + SESSION_TTL_SECONDS
    return session


def google_connection_status() -> dict[str, object]:
    connection = load_google_connection()
    return {
        "googleConfigured": google_oauth_configured(),
        "googleConnected": google_connected(),
        "googleEmail": str(connection.get("google_email", "")).strip(),
        "spreadsheetId": str(connection.get("spreadsheet_id", "")).strip(),
        "spreadsheetUrl": str(connection.get("spreadsheet_url", "")).strip(),
    }


class InvitelyHandler(BaseHTTPRequestHandler):
    server_version = "InvitelyServer/2.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path in STATIC_ROUTES:
            self.serve_file(BASE_DIR / STATIC_ROUTES[parsed.path])
            return

        if parsed.path == "/api/events":
            self.send_json({"events": get_events()})
            return

        if parsed.path == "/api/settings":
            self.send_json({"settings": get_settings()})
            return

        if parsed.path == "/api/admin/status":
            session = get_admin_session(self)
            payload = {"authenticated": bool(session)}
            payload.update(google_connection_status())
            self.send_json(payload)
            return

        if parsed.path == "/api/admin/rsvps":
            self.require_admin_session()
            self.send_json({"rsvps": get_rsvps()})
            return

        if parsed.path == "/auth/google/start":
            self.handle_google_start()
            return

        if parsed.path == "/auth/google/callback":
            self.handle_google_callback(parsed)
            return

        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)

        try:
            payload = self.read_json()
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            if parsed.path == "/api/rsvp":
                add_rsvp(payload)
                self.send_json({"ok": True})
                return

            if parsed.path == "/api/admin/session":
                password = str(payload.get("password", "")).strip()
                if password != ADMIN_PASSWORD:
                    self.send_json({"error": "Host password is invalid."}, status=HTTPStatus.UNAUTHORIZED)
                    return
                session_token = create_admin_session()
                self.send_json(
                    {"ok": True},
                    cookies=[self.build_session_cookie(session_token)],
                )
                return

            self.require_admin_session()

            if parsed.path == "/api/admin/events":
                upsert_event(payload)
                self.send_json({"ok": True})
                return

            if parsed.path == "/api/admin/settings":
                save_settings(payload)
                self.send_json({"ok": True})
                return

            if parsed.path == "/api/admin/events/delete":
                delete_event(str(payload.get("id", "")).strip())
                self.send_json({"ok": True})
                return
        except PermissionError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.UNAUTHORIZED)
            return
        except ValueError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return
        except RuntimeError as error:
            self.send_json({"error": str(error)}, status=HTTPStatus.BAD_REQUEST)
            return

        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def handle_google_start(self) -> None:
        session = self.require_admin_session()
        if not google_oauth_configured():
            self.send_redirect("/?google=failed&error=" + quote("Google OAuth is not configured on the server."))
            return

        _, _, Flow, _ = import_google_libraries()
        redirect_uri = get_google_redirect_uri(self)
        flow = Flow.from_client_config(google_client_config(self), scopes=GOOGLE_SCOPES, redirect_uri=redirect_uri)
        auth_url, state = flow.authorization_url(
            access_type="offline",
            include_granted_scopes="true",
            prompt="consent",
        )
        session["google_oauth_state"] = state
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", auth_url)
        self.end_headers()

    def handle_google_callback(self, parsed) -> None:
        session = self.require_admin_session()
        params = parse_qs(parsed.query)
        error = params.get("error", [""])[0]
        if error:
            self.send_redirect("/?google=failed&error=" + quote(error))
            return

        state = params.get("state", [""])[0]
        code = params.get("code", [""])[0]
        expected_state = str(session.get("google_oauth_state", ""))

        if not code or not state or state != expected_state:
            self.send_redirect("/?google=failed&error=" + quote("Invalid Google OAuth state."))
            return

        _, _, Flow, _ = import_google_libraries()
        redirect_uri = get_google_redirect_uri(self)
        flow = Flow.from_client_config(google_client_config(self), scopes=GOOGLE_SCOPES, redirect_uri=redirect_uri)
        flow.fetch_token(code=code)
        credentials = flow.credentials
        oauth_api = google_service("oauth2", "v2", credentials)
        user_info = oauth_api.userinfo().get().execute()

        connection = load_google_connection()
        spreadsheet_id = str(connection.get("spreadsheet_id", "")).strip()
        spreadsheet_url = str(connection.get("spreadsheet_url", "")).strip()
        created_new_sheet = False

        if not spreadsheet_id:
            settings = workbook_get_settings()
            sheet_info = google_create_spreadsheet(f"{settings.get('appName', 'Invitely')} Data", credentials)
            spreadsheet_id = sheet_info["spreadsheet_id"]
            spreadsheet_url = sheet_info["spreadsheet_url"]
            created_new_sheet = True

        save_google_connection(
            {
                "refresh_token": credentials.refresh_token or str(connection.get("refresh_token", "")).strip(),
                "spreadsheet_id": spreadsheet_id,
                "spreadsheet_url": spreadsheet_url,
                "google_email": str(user_info.get("email", "")).strip(),
                "connected_at": datetime.now().isoformat(),
            }
        )

        if created_new_sheet:
            migrate_workbook_to_google()

        session["google_oauth_state"] = ""
        self.send_redirect("/?google=connected")

    def serve_file(self, path: Path) -> None:
        if not path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        if path.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif path.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif path.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        else:
            content_type = "application/octet-stream"

        body = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict[str, object]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        raw_body = self.rfile.read(length)
        try:
            return json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError as error:
            raise ValueError("Invalid JSON payload.") from error

    def build_session_cookie(self, session_token: str) -> str:
        cookie = SimpleCookie()
        cookie[SESSION_COOKIE_NAME] = session_token
        cookie[SESSION_COOKIE_NAME]["path"] = "/"
        cookie[SESSION_COOKIE_NAME]["httponly"] = True
        cookie[SESSION_COOKIE_NAME]["samesite"] = "Lax"
        return cookie.output(header="").strip()

    def require_admin_session(self) -> dict[str, object]:
        session = get_admin_session(self)
        if not session:
            raise PermissionError("Host session is invalid. Please sign in again.")
        return session

    def send_redirect(self, location: str) -> None:
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", location)
        self.end_headers()

    def send_json(
        self,
        payload: dict[str, object],
        status: HTTPStatus = HTTPStatus.OK,
        cookies: list[str] | None = None,
    ) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for cookie in cookies or []:
            self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format: str, *args: object) -> None:
        return


def main() -> None:
    ensure_workbook()
    server = ThreadingHTTPServer((SERVER_HOST, SERVER_PORT), InvitelyHandler)
    print(f"Invitely is running on http://{SERVER_HOST}:{SERVER_PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
